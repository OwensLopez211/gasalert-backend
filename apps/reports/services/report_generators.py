from io import BytesIO
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from reportlab.lib.utils import ImageReader
from reportlab.lib.pagesizes import letter
import csv
import logging

logger = logging.getLogger(__name__)

def generate_pdf_report(data):
    """
    Genera un reporte en formato PDF basado en los datos proporcionados.
    """
    try:
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)

        # Configuración inicial del PDF
        pdf.setTitle("Reporte de Tanques")
        width, height = letter
        y_position = height - 50  # Margen inicial superior

        # Título del Reporte
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawString(50, y_position, "Reporte de Estado de Tanques")
        y_position -= 30

        # Fecha y rango del reporte
        pdf.setFont("Helvetica", 10)
        pdf.drawString(50, y_position, f"Generado el: {data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}")
        pdf.drawString(50, y_position - 15, f"Rango de datos: {data['date_range']}")
        y_position -= 40

        # Estado de los Tanques
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y_position, "Estado de Tanques")
        y_position -= 20

        pdf.setFont("Helvetica", 10)
        for tank in data.get('tanks_status', []):
            pdf.drawString(50, y_position, f"Tanque: {tank['nombre']}")
            pdf.drawString(200, y_position, f"Nivel: {tank.get('nivel_actual', 0)}%")
            pdf.drawString(300, y_position, f"Estado: {tank.get('estado', 'N/A')}")
            y_position -= 15
            if y_position < 50:
                pdf.showPage()
                y_position = height - 50

        # Resumen de Alertas
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y_position, "Resumen de Alertas")
        y_position -= 20

        pdf.setFont("Helvetica", 10)
        alerts_data = data.get('alerts', {})
        pdf.drawString(50, y_position, f"Total de alertas: {alerts_data.get('total', 0)}")
        y_position -= 15

        for alert_type, count in alerts_data.get('by_type', {}).items():
            pdf.drawString(50, y_position, f"- {alert_type}: {count}")
            y_position -= 15

        # Métricas de Consumo
        y_position -= 20
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y_position, "Métricas de Consumo")
        y_position -= 20

        pdf.setFont("Helvetica", 10)
        consumption_metrics = data.get('consumption_metrics', {})
        pdf.drawString(50, y_position, f"Consumo Total: {consumption_metrics.get('total_consumption', 0)} litros")
        y_position -= 15
        pdf.drawString(50, y_position, f"Consumo Promedio Diario: {consumption_metrics.get('daily_avg', 0)} litros")
        y_position -= 20

        # Inserción de Gráfico
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y_position, "Tendencias de Consumo Diario")
        y_position -= 200

        consumption_graph = data['graphs'].get('consumption_graph')
        if consumption_graph:
            graph_image = ImageReader(consumption_graph)
            pdf.drawImage(graph_image, 50, y_position, width=500, height=200)
            y_position -= 220

        # Finalización
        pdf.showPage()
        pdf.save()
        buffer.seek(0)

        return buffer
    except Exception as e:
        raise Exception(f"Error generando PDF: {e}")


def generate_excel_report(data):
    """
    Genera un reporte en formato Excel basado en los datos proporcionados.
    """
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Tanques"

        # Agregar encabezados
        headers = ["Tanque", "Consumo Diario", "Consumo Semanal", "Nivel Actual", "Estado"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        # Agregar datos
        for row, tank in enumerate(data.get('tanks_status', []), 2):
            sheet.cell(row=row, column=1, value=tank.get('nombre', ''))
            sheet.cell(row=row, column=2, value=tank.get('consumo_diario', 0))
            sheet.cell(row=row, column=3, value=tank.get('consumo_semanal', 0))
            sheet.cell(row=row, column=4, value=tank.get('nivel_actual', 0))
            sheet.cell(row=row, column=5, value=tank.get('estado', 'N/A'))

        # Ajustar ancho de columnas
        for column in sheet.columns:
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            sheet.column_dimensions[column[0].column_letter].width = max_length + 2

        # Preparar el archivo para la descarga
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        workbook.save(response)
        logger.info("Excel generado exitosamente.")
        return response
    except Exception as e:
        logger.error(f"Error generando Excel: {e}", exc_info=True)
        raise

def generate_csv_report(data):
    """
    Genera un reporte en formato CSV basado en los datos proporcionados.
    """
    try:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reporte_tanques.csv"'

        writer = csv.writer(response)
        headers = ["Tanque", "Consumo Diario", "Consumo Semanal", "Nivel Actual", "Estado"]
        writer.writerow(headers)

        for tank in data.get('tanks_status', []):
            writer.writerow([
                tank.get('nombre', ''),
                tank.get('consumo_diario', 0),
                tank.get('consumo_semanal', 0),
                tank.get('nivel_actual', 0),
                tank.get('estado', 'N/A')
            ])

        logger.info("CSV generado exitosamente.")
        return response
    except Exception as e:
        logger.error(f"Error generando CSV: {e}", exc_info=True)
        raise
